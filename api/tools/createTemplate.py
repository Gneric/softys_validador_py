from email.quoprimime import header_check
import json
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter

def createErrorTemplate(file, data):
    try:
        filename = str(file.filename).replace('.xlsx','-Errors.xlsx')
        filepath = f'files/{filename}'
        df = pd.read_excel(file, sheet_name='data')
        cols = df.columns
        # Delete if exists
        if os.path.exists(filepath):
            print(f'Deleting {filename}')
            os.remove(filepath)
            
        error_list = [ { 
            'index': row.get('row'), 
            'columns': [ c.get('column') for c in row.get('error_details') ]  
            } for row in json.loads(data) ]
        with pd.ExcelWriter(filepath) as writer:
            df.to_excel(writer, sheet_name='data', index=False, )

        pttrn = PatternFill("solid", fgColor="FA2222")
        print(error_list)
        wb = load_workbook(filename = filepath)
        ws = wb.active
        for error in error_list:
          for row in ws.iter_rows(2):
            if str(row[0].row) == str(error['index']):
              for cell in row:
                if cols[row.index(cell)] in error['columns']:
                    cell.fill = pttrn
                
        wb.save(filepath)            

        return filename

    except:
        print(f'error in createErrorTemplate {sys.exc_info()}')
        return ''

def createTemplate(structure):
    try:
        data = json.loads(structure[0][0])
        validata = data.get('validationData')
        processName = data.get("processName")
        processID = validata[0].get('processID')
        filename = f'{processID}_{processName}.xlsx'
        filepath = f'files/{filename}'

        # Delete if exists
        if os.path.exists(filepath):
            print(f'Deleting {filename}')
            os.remove(filepath)

        workbook = xlsxwriter.Workbook(filepath, {'in_memory': True})
        data_ws = workbook.add_worksheet('data')
        # Formato string
        text_format = workbook.add_format()
        text_format.set_num_format('@')
        # Formato int
        int_format = workbook.add_format()
        int_format.set_num_format('#')
        # Formato float
        float_format = workbook.add_format()
        float_format.set_num_format('#,##0.00')
        # Formato date
        date_format = workbook.add_format()
        date_format.set_num_format('d/mm/yyyy')

        format_selector = {
            'string': text_format,
            'int': int_format,
            'float': float_format,
            'date': date_format
        }

        header_format = workbook.add_format()
        #header_format.set_pattern(1)
        header_format.set_bg_color('4682b4')
        header_format.set_bold()
        header_format.set_text_wrap()
        
        

        for row in validata:
            data_ws.write(0, validata.index(row), row.get('columnName'), header_format)
            data_ws.write_comment('A1', 'This is a comment')
            data_ws.set_column( 
                first_col=validata.index(row), 
                last_col=validata.index(row), 
                width=20, 
                cell_format=format_selector.get(row.get('columnType'))
            )

        
        workbook.close()
        return filename
    except Exception as err:
        print('Error in createTemplate :', err)
        return ''